import os
import openpyxl
import random
from crud.invite_code import create_invite_code
from crud.club import create_club, get_club_by_name
from database import SessionLocal

os.environ.setdefault('DATABASE_URL', 'sqlite:////www/bupt-bingo/Server/bingo.db')

def generate_random_strings(file_path):
    db = SessionLocal()
    # 加载xlsx文件（支持写入）
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 获取第一个工作表
    
    # 遍历行（从第二行开始，min_row=2）
    # 注意：iter_rows返回的row是单元格元组，row[0]是第一列，row[1]是第二列...
    for row in sheet.iter_rows(min_row=2):
        print(row[0].value, row[1].value)  # 打印第一列和第二列内容
        # 获取当前行的行号（用于准确定位第三列）
        row_num = row[0].row  # 取第一列单元格的行号（确保行号正确）
        
        # 获取第二列内容（索引1），若第二列不存在则跳过
        if len(row) < 2:  # 检查当前行是否至少有2列
            continue
        col2_value = row[1].value
        if not col2_value:
            continue
        
        # 创建社团
        if get_club_by_name(db, club_name=col2_value):
            continue  # 社团已存在则跳过
        club = create_club(db, club_name=col2_value, club_type=0)
        print(f"创建社团: id={club.id}, name={club.club_name}")
        
        # 生成3个邀请码
        row_invite_codes = []
        for _ in range(3):
            random_num = random.randint(0, 999)
            random_str = f"{col2_value}{random_num:03d}"
            row_invite_codes.append(random_str)
            create_invite_code(db, code=random_str, role=1, club_id=club.id)
        
        # 关键修改：通过行号+列号定位第三列（column=3，Excel列A=1，B=2，C=3）
        # 即使原行没有第三列，也会自动创建单元格
        third_col_cell = sheet.cell(row=row_num, column=3)
        third_col_cell.value = ",".join(row_invite_codes)
    
    # 生成管理员邀请码（不写入Excel）
    for num in random.sample(range(100, 1000), 10):
        random_str = f"admin{num:03d}"
        create_invite_code(db, code=random_str, role=2, club_id=None)
        print(f"生成管理员邀请码: {random_str}")
    
    # 保存并关闭
    workbook.save(file_path)
    workbook.close()
    db.close()

if __name__ == "__main__":
    file_path = "附件3 参与社团名单.xlsx"
    generate_random_strings(file_path)